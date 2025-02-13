import requests
import openpyxl
import json
import os
from datetime import datetime

api_url = "https://avayzaravand.ir/wp-json/attendance/v1/upload"
excel_file = r'C:\Users\MS\Desktop\attendance_data.xlsx'

def validate_sheet_name(sheet_name):
    """اعتبارسنجی فرمت تاریخ در نام شیت (1403-04-19)"""
    try:
        datetime.strptime(sheet_name, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def get_latest_valid_sheet(wb):
    """یافتن آخرین شیت با فرمت تاریخ معتبر"""
    valid_sheets = [s for s in wb.sheetnames if validate_sheet_name(s)]
    if not valid_sheets:
        raise ValueError("هیچ شیت معتبری با فرمت تاریخ یافت نشد")
    return sorted(valid_sheets, key=lambda x: datetime.strptime(x, '%Y-%m-%d'))[-1]

def send_data():
    try:
        # بررسی وجود فایل
        if not os.path.isfile(excel_file):
            raise FileNotFoundError(f"فایل {excel_file} یافت نشد")
            
        wb = openpyxl.load_workbook(excel_file, read_only=True)
        
        # انتخاب آخرین شیت معتبر
        last_sheet_name = get_latest_valid_sheet(wb)
        ws = wb[last_sheet_name]
        
        # پردازش داده‌ها
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue
                
            user_id, persian_name, user_class, attendance_status = row[:4]
            
            record = {
                "user_id": int(user_id),
                "persian_name": persian_name.strip(),
                "user_class": user_class.strip(),
                "attendance_status": attendance_status.strip().replace('غیبت', 'غیاب'),
                "date_sheet": last_sheet_name
            }
            
            # اعتبارسنجی نهایی
            if all(record.values()) and record['user_id'] > 0:
                data.append(record)

        # ارسال به سرور
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'Basic ' + os.getenv('API_KEY', '')  # اختیاری
        }
        
        response = requests.post(
            api_url,
            json=data,
            headers=headers,
            timeout=10
        )

        response.raise_for_status()
        
        print(f"✅ موفق! داده‌های شیت {last_sheet_name} ارسال شدند")
        print(f"🆔 کد پاسخ: {response.status_code}")
        print(f"📨 پاسخ سرور: {response.json()}")

    except Exception as e:
        print(f"❌ خطا: {str(e)}")
        if hasattr(e, 'response') and e.response:
            print(f"🔧 جزئیات خطا: {e.response.text}")

if __name__ == "__main__":
    send_data()