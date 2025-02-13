from zk import ZK, const
import openpyxl
from khayyam import JalaliDatetime
from datetime import datetime
from collections import defaultdict

# Device connection details
device_ip = '192.168.1.146'
port = 4370

# Excel file to save the data
excel_file = 'attendance_data.xlsx'

# Path to the Excel file containing User ID, Persian Name, and Class
user_info_file = 'user_info.xlsx'

try:
    # Step 1: Read the user info Excel file
    user_info_wb = openpyxl.load_workbook(user_info_file)
    user_info_ws = user_info_wb.active

    # Create dictionaries to map User ID to Persian Name and Class
    user_id_to_name = {}
    user_id_to_class = {}
    for row in user_info_ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        user_id, persian_name, user_class = row[0], row[1], row[2]
        user_id_to_name[int(user_id)] = persian_name  # تبدیل User ID به عددی
        user_id_to_class[int(user_id)] = user_class  # ذخیره کلاس

    print(f"Loaded {len(user_id_to_name)} user mappings from {user_info_file}.")

    # Step 2: Connect to the device
    zk = ZK(device_ip, port=port, timeout=5, password=0, force_udp=False, ommit_ping=False)
    conn = zk.connect()
    conn.disable_device()  # Disable the device to avoid interruptions

    print("Connection established, and the device is disabled.")
    
    # Step 3: Fetch attendance data
    attendance = conn.get_attendance()

    if attendance:
        print(f"{len(attendance)} attendance records found.")
        
        # Create or open an Excel file
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:  # Remove default sheet
            wb.remove(wb['Sheet'])
        
        # Group attendance records by date
        attendance_by_date = defaultdict(list)
        for record in attendance:
            record_date = record.timestamp.date()  # Extract date from timestamp
            attendance_by_date[record_date].append(record)
        
        # Process each day's attendance
        for date, records in attendance_by_date.items():
            # Convert Gregorian date to Shamsi
            shamsi_date = JalaliDatetime(date).strftime('%Y-%m-%d')  # Format as Shamsi date
            
            # Create a new sheet for the day with Shamsi date as the name
            sheet_name = shamsi_date
            ws = wb.create_sheet(title=sheet_name)
            
            # Copy user info to the new sheet
            ws.append(['User ID', 'Persian Name', 'Class', 'حضور و غیاب'])  # Add headers
            for row in user_info_ws.iter_rows(min_row=2, values_only=True):  # Skip header row
                user_id, persian_name, user_class = row[0], row[1], row[2]
                ws.append([user_id, persian_name, user_class, 'غیاب'])  # Default to 'غیاب'
            
            # Mark attendance for the day
            for record in records:
                user_id = int(record.user_id)  # تبدیل User ID به عددی
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Check User ID column
                    if row[0] == user_id:
                        ws.cell(row=row_idx, column=4, value='حاضر')  # Update حضور و غیاب
                        break
        
        # Save the Excel file
        wb.save(excel_file)
        print(f"Data successfully saved in the file {excel_file}.")
    else:
        print("No attendance data found.")

except Exception as e:
    print(f"Error: {e}")

finally:
    # Ensure the device is enabled and disconnected
    if conn:
        conn.enable_device()  # Enable the device again
        conn.disconnect()
        print("Device enabled and disconnected.")